import json
from openpyxl import load_workbook
wb = load_workbook("/home/pi/simple_test/pythons/Copy of Astoria-SIT-TestPlan_BMCQA_DVT.xlsx")
ws = wb["OEM IPMI Cmds"]
f = open('/home/pi/simple_test/ipmi_table.json')
ipmi_tb = json.load(f)
for i in range(1,50):
    val=ws[f'D{i}'].value
    if val !=None:
        cmds=ws[f'L{i}'].value
        if cmds!= None:
            cmdList=cmds.split('\n')
            oList=[]
            for c in cmdList:
                if 'ipmitool' in c:
                    if '#' in c:
                        oList.append(c[c.index('#')+1:].strip())
                    else:
                        oList.append(c.strip())
            ipmi_tb[val]=oList
with open('/home/pi/simple_test/ipmi_table.json', "w+") as f:
    json.dump(ipmi_tb, f, indent = 4)