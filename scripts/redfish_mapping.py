import json
from openpyxl import load_workbook
wb = load_workbook("/home/pi/simple_test/pythons/Copy of Astoria-SIT-TestPlan_BMCQA_DVT.xlsx")
ws = wb["Redfish check"]
redfish_tb = {}
for i in range(12,150):
    val=ws[f'D{i}'].value
    if val !=None:
        cmds=ws[f'F{i}'].value
        if cmds!= None:
            cmdList=cmds.split('\n')
            for c in cmdList:
                if 'curl' in c:
                    redfish_tb[val]=c.strip()
                    break
with open('/home/pi/simple_test/redfish_table.json', "w+") as f:
    json.dump(redfish_tb, f, indent = 4)