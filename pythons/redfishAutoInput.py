import json,subprocess
from openpyxl import load_workbook
from robot.libraries.BuiltIn import BuiltIn
oldHost='2001:db8:0:1:1622:3bff:feaa:fa58'
newHost=BuiltIn().get_variable_value("${Redfish_Host}")
f = open(f'{BuiltIn().get_variable_value("${PYTHONS_PATH}")}/redfish_table.json')
redfish_tb = json.load(f)
model_xlsx = f'{BuiltIn().get_variable_value("${RESULT_PATH}")}/sdrAutoInput.xlsx'
#model_xlsx=BuiltIn().get_variable_value("${XLSX_MODEL}")
output_xlsx = f'{BuiltIn().get_variable_value("${RESULT_PATH}")}/redfishAutoInput.xlsx'

def execute_command(command):
    while True:
        sp = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE)
        subprocess_return = sp.stdout.read().decode()
        if subprocess_return != None and subprocess_return != '':
            return f"{command}\n{subprocess_return}\n"

def redfish_autoinput():
    wb = load_workbook(model_xlsx)
    ws = wb["Redfish check"]
    for i in range(12,150):
        val=ws[f'D{i}'].value
        if val !=None:
            cmd=redfish_tb.get(val)
            if cmd != None:
                ws[f'F{i}']=execute_command(cmd.replace(oldHost,newHost))
    wb.save(output_xlsx)
