import json
from openpyxl import load_workbook
from robot.libraries.BuiltIn import BuiltIn
curpath = BuiltIn().get_variable_value("${PYTHONS_PATH}") + '/'
resPath = BuiltIn().get_variable_value("${RESULT_PATH}") + '/'
thresholdPath = resPath + BuiltIn().get_variable_value("${THRESHOLD}") + '.txt'
sdrinfoPath = resPath + BuiltIn().get_variable_value("${SDRINFO}") + '.txt'
specPath = BuiltIn().get_variable_value("${SPEC_PATH}")
sdrlistPath = f'{BuiltIn().get_variable_value("${SANITY_PATH}")}/sensor_list_all.txt'
hexHead = '0x'
threParams = ['Lower Critical','Lower Non-Critical','Upper Non-Critical','Upper Critical']
sdrParams = [
'entity id is',
'entity is',
'sensor type is',
'reading type is'
]
def dumpJson(jsonName,dictVar):
    with open(curpath + jsonName, "w+") as f:
        json.dump(dictVar, f, indent = 4)
def sdrlistMapping():
    with open(sdrlistPath) as f:
        sdrlistLines = f.readlines()
    thresholdMap = thresholdMapping()
    for line in sdrlistLines:
        if 'ipmitool' in line:
            continue
        sdrs = line.split('|')
        sensorName = sdrs[0].strip().replace(" ", "_")
        if thresholdMap.get(sensorName) != None:
            thresholdMap[sensorName]['Sensor Reading'] = float(sdrs[1].strip()) if sdrs[1].strip()!='na' else None
            thresholdMap[sensorName]['Lower Critical'] = float(sdrs[-5].strip()) if sdrs[-5].strip()!='na' else None
            thresholdMap[sensorName]['Lower Non-Critical'] = float(sdrs[-4].strip()) if sdrs[-4].strip()!='na' else None
            thresholdMap[sensorName]['Upper Non-Critical'] = float(sdrs[-3].strip()) if sdrs[-3].strip()!='na' else None
            thresholdMap[sensorName]['Upper Critical'] = float(sdrs[-2].strip()) if sdrs[-2].strip()!='na' else None
            thresholdMap[sensorName]['sdrListResult'] = line
    #dumpJson("threshold.json",thresholdMap)
    return thresholdMap
def sdrinfoMapping():
    with open(sdrinfoPath) as f:
        sdrinfoLines = f.readlines()
    sdrinfoMap = {}
    tmpHex = ''
    count = 0
    tmpIndex = 0
    for line in sdrinfoLines:
        if 'ipmitool raw' in line:
            tmpHex = line[33:37]
            sdrinfoMap[tmpHex] = {'Record Type':hexHead+line[-4:-2].upper()}
        if 'sensor ID is' in line or count == len(sdrinfoLines)-1:
            if count > 5:
                sdrResult = ''
                for i in range(tmpIndex,count):
                    if '|' not in sdrinfoLines[i] and 'sensor ID is'  not in sdrinfoLines[i]:
                        sdrResult+=sdrinfoLines[i]#+'\n'
                sdrinfoMap[tmpHex]['sdrResult'] = sdrResult
            tmpIndex = count
        for sp in sdrParams:
            if sp in line:
                if sp[:-3] == 'entity':
                    sdrinfoMap[tmpHex][sp[:-3]] = str(int(line[-4:-2],16))
                else:
                    sdrinfoMap[tmpHex][sp[:-3]] = hexHead+line[-4:-2].upper()
        count+=1
    #dumpJson("sdrinfo.json",sdrinfoMap)
    return sdrinfoMap

def thresholdMapping():
    with open(thresholdPath) as f:
        thresholdLines = f.readlines()
    thresholdMap = {}
    tmpName = ''
    count = 0
    for line in thresholdLines:
        if 'Sensor ID' in line:
            if count > 5:
                thresResult = ''
                for i in range(tmpIndex,count-3):
                    thresResult+=thresholdLines[i]
                    if i == tmpIndex:
                        thresResult += '\n'
                thresholdMap[tmpName]['thresResult'] = thresResult
            tmpIndex = count
            subIndex1 = line.find(':')
            subIndex2 = line.find('(')
            subIndex3 = line.find(')')
            sensorName = line[subIndex1+2:subIndex2-1].replace(" ", "_")
            sensorHex = line[subIndex2+3:subIndex3]
            if len(sensorHex) < 2:
                sensorHex = '0'+sensorHex
            thresholdMap[sensorName]={'hex':hexHead+sensorHex.upper()}
            tmpName = sensorName
        for tp in threParams:
            if tp in line:
                subIndex = line.find(':')+2
                if line[subIndex:] == 'na\n':
                    thresholdMap[tmpName][tp] = None
                else:
                    if line[subIndex:-1] not in threParams:
                        thresholdMap[tmpName][tp] = float(line[subIndex:-1])
        count+=1

    #dumpJson("threshold.json",thresholdMap)
    return thresholdMap

def sdrMapping():
    thresholdMap = sdrlistMapping()
    sdrinfoMap = sdrinfoMapping()

    for tKey in thresholdMap:
        if sdrinfoMap.get(thresholdMap[tKey].get('hex'))==None:
            continue
        thresholdMap[tKey].update(sdrinfoMap.get(thresholdMap[tKey].get('hex')))

    dumpJson("sdrTable.json",thresholdMap)
    return thresholdMap

def specMapping():
    wb = load_workbook(specPath)
    ws = wb["SDR Table"]
    table = {
        'entID':{
            'col':'F',
            'key':'entity id'
        },
        'ent':{
            'col':'G',
            'key':'entity'
        },
        'UC':{
            'col':'O',
            'key':'Upper Critical'
        },
        'UN':{
            'col':'P',
            'key':'Upper Non-Critical'
        },
        'LN':{
            'col':'Q',
            'key':'Lower Non-Critical'
        },
        'LC':{
            'col':'R',
            'key':'Lower Critical'
        }
    }
    specMap={}
    for i in range(3,200):
        sensorName = ws['B'+str(i)].value
        if sensorName != None and sensorName != "":
            specMap[sensorName]={}
            for k in table.keys():
                val=ws[table[k]['col']+str(i)].value
                if val != None:
                    if isinstance(val, str):
                        if hexHead in val and len(val)==3:
                            val=f"{hexHead}0{val[-1]}"
                        if val == "-" or val == "" or val == "NA" or val == "na":
                            val=None
                    if k == 'ent':
                        val=str(int(val))
                specMap[sensorName][table[k]['key']]=val
    dumpJson("spec.json",specMap)

def sdrTable_mapping():
    sdrMapping()
    specMapping()
