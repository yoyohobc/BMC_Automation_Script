import paramiko,json
from openpyxl import load_workbook
from robot.libraries.BuiltIn import BuiltIn
f = open(f'{BuiltIn().get_variable_value("${PYTHONS_PATH}")}/ipmi_table.json')
ipmi_tb = json.load(f)
model_xlsx = f'{BuiltIn().get_variable_value("${RESULT_PATH}")}/redfishAutoInput.xlsx'
#model_xlsx=BuiltIn().get_variable_value("${XLSX_MODEL}")
output_xlsx = f'{BuiltIn().get_variable_value("${RESULT_PATH}")}/AutoInput.xlsx'
ssh = paramiko.SSHClient()
ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
host=BuiltIn().get_variable_value("${BMC_HOST}")
username=BuiltIn().get_variable_value("${BMC_USER}")
password=BuiltIn().get_variable_value("${BMC_PASSWORD}")
ssh.connect(host, username=username, password=password)

def execute_command(command):
    while True:
        try:
            ssh_stdin, ssh_stdout, ssh_stderr = ssh.exec_command(command)
            while True:
                try:
                    output = ssh_stdout.read()
                    break
                except:
                    continue
            return f"{command}\n{output.decode()}\n"
        except:
            ssh.connect(host, username=username, password=password)

def tab_input(ws):
    for i in range(1,50):
        val=ws[f'D{i}'].value
        if val !=None:
            if ipmi_tb.get(val) != None:
                result=''
                for cmd in ipmi_tb.get(val):
                    result+=execute_command(cmd)
                ws[f'L{i}']=result
    return ws

def ipmi_autoinput():
    wb = load_workbook(model_xlsx)
    wc = wb["OEM IPMI Cmds"]
    wd = wb["IPMI Cmds"]
    wc=tab_input(wc)
    wd=tab_input(wd)
    wb.save(output_xlsx)